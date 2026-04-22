from __future__ import annotations

import configparser
import difflib
import json
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import os

# 修订记录（按时间顺序）
# V1: 基础版，支持固定路径读取模板与输入文件并写出结果
# V2: 增加字段映射与“接收日期”双写入（创建时间、回复时间/ITM回函时间）
# V3: 增加按“工单编号/ITM编号”去重与模板列对齐
# V4: 增加客户类型识别、客户名称前缀清理、回复人映射
# V5: 增加日期列解析与 Excel 日期格式/列宽设置
# V6: 增加 ini 配置化输入（TEMPLATE_PATH、INPUT_PATHS）并支持多文件
# V7: 增强容错（相似文件名匹配、缺失输入跳过）并补全回复人映射规则
# V8: 固定配置文件名为 merge_requirement_to_template.ini，
#     并将标题映射/回复人映射统一迁移到 ini（Title_Map/TaskAssigner_Map）
# V9: 增加运行日志文件输出（新增数、去重后总数、重复工单编号清单）
# V10: 增加路径存在性预检查与友好报错建议，日志同时写入文件与标准输出
# V11: 新增数据按“创建时间”排序后追加到模板末尾，避免与原始模板数据混编
# V12: 输出 Excel 打开时默认定位到“原模板数据倒数5行”位置，便于续看历史数据
# V13: 读取 ini 使用 utf-8-sig，兼容 Windows 记事本保存的带 BOM 的 UTF-8
# 解决了pyinstaller 打包以后找不到配置文件的问题
# V14：20260408-给“工单状态”默认赋值“进行中”，“事项来源”默认赋值“ITM”
CONFIG_PATH = Path(__file__).with_name("merge_requirement_to_template.ini")
CONFIG_SECTION = "PATHS"

TEMPLATE_SHEET = "模板页"


def normalize(s: object) -> str:
    if s is None or pd.isna(s):
        return ""
    return str(s).strip()


def build_output_path(template_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return template_path.with_name(f"{template_path.stem}_{ts}{template_path.suffix}")


def build_log_path(script_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return script_path.with_name(f"{script_path.stem}_{ts}.log")


def parse_input_paths(raw_value: str) -> list[Path]:
    """
    INPUT_PATHS supports:
    - one path per line
    - comma-separated paths
    - mixed line/comma style
    """
    normalized = raw_value.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("，", ",").replace("；", ";")
    parts: list[str] = []
    for line in normalized.split("\n"):
        semicolon_split = []
        for block in line.split(";"):
            semicolon_split.extend(block.split(","))
        for piece in semicolon_split:
            candidate = piece.strip().strip('"').strip("'")
            if candidate:
                parts.append(candidate)
    return [Path(p) for p in parts]


def normalize_filename_key(name: str) -> str:
    return (
        name.strip()
        .replace("（", "(")
        .replace("）", ")")
        .replace("【", "[")
        .replace("】", "]")
        .replace("，", ",")
        .replace("。", ".")
        .replace("：", ":")
        .replace(" ", "")
        .lower()
    )


def resolve_similar_excel_path(path: Path) -> Path:
    if path.exists():
        return path
    parent = path.parent
    if not parent.exists():
        return path

    target_key = normalize_filename_key(path.name)
    candidates = list(parent.glob("*.xlsx")) + list(parent.glob("*.xlsm"))
    for candidate in candidates:
        if normalize_filename_key(candidate.name) == target_key:
            return candidate
    return path


def build_path_hint(path: Path) -> str:
    parent = path.parent
    if not parent.exists():
        return (
            f"路径不存在: {path}\n"
            f"建议: 先确认目录是否存在 -> {parent}\n"
            "建议: 检查是否把中文括号/英文括号写错，或是否缺少盘符前缀。"
        )

    excel_candidates = [p.name for p in parent.iterdir() if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]
    if not excel_candidates:
        return (
            f"路径不存在: {path}\n"
            f"建议: 目录存在但未发现 Excel 文件 -> {parent}\n"
            "建议: 检查文件是否尚未保存、后缀是否正确。"
        )

    matches = difflib.get_close_matches(path.name, excel_candidates, n=3, cutoff=0.45)
    if matches:
        candidates_text = "\n".join(f"  - {name}" for name in matches)
        return (
            f"路径不存在: {path}\n"
            "建议: 发现同目录相似文件名，可能是括号/标点/全半角差异：\n"
            f"{candidates_text}"
        )

    sample = "\n".join(f"  - {name}" for name in excel_candidates[:5])
    return (
        f"路径不存在: {path}\n"
        "建议: 同目录存在以下 Excel 文件，请核对是否应使用其中之一：\n"
        f"{sample}"
    )


def parse_title_map(raw_value: str) -> dict[str, list[str]]:
    if not raw_value.strip():
        raise ValueError("配置项 Title_Map 不能为空")
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Title_Map 不是合法 JSON: {exc}") from exc

    if not isinstance(parsed, dict):
        raise ValueError("Title_Map 必须是 JSON 对象（键值映射）")

    normalized_map: dict[str, list[str]] = {}
    for src_col, target in parsed.items():
        src_name = normalize(src_col)
        if src_name == "":
            continue
        if isinstance(target, str):
            target_names = [normalize(target)]
        elif isinstance(target, list):
            target_names = [normalize(x) for x in target if isinstance(x, str)]
        else:
            raise ValueError(f"Title_Map 中字段 {src_col} 的映射值必须是字符串或字符串数组")

        target_names = [x for x in target_names if x != ""]
        if not target_names:
            continue
        normalized_map[src_name] = target_names

    if not normalized_map:
        raise ValueError("Title_Map 解析后为空，请检查配置")
    return normalized_map


def parse_task_assigner_map(raw_value: str) -> dict[str, str]:
    if not raw_value.strip():
        raise ValueError("配置项 TaskAssigner_Map 不能为空")
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"TaskAssigner_Map 不是合法 JSON: {exc}") from exc

    if not isinstance(parsed, dict):
        raise ValueError("TaskAssigner_Map 必须是 JSON 对象（键值映射）")

    normalized_map: dict[str, str] = {}
    for org_name, assignee in parsed.items():
        org_key = normalize(org_name)
        assignee_name = normalize(assignee)
        if org_key and assignee_name:
            normalized_map[org_key] = assignee_name

    if not normalized_map:
        raise ValueError("TaskAssigner_Map 解析后为空，请检查配置")
    return normalized_map


def load_config_from_ini(
    config_path: Path,
) -> tuple[Path, list[Path], list[Path], dict[str, list[str]], dict[str, str]]:
    if not config_path.exists():
        raise FileNotFoundError(
            f"未找到配置文件: {config_path}\n"
            "请在脚本同目录创建 merge_requirement_to_template.ini，"
            "并在 [PATHS] 中配置 TEMPLATE_PATH、INPUT_PATHS、Title_Map、TaskAssigner_Map。"
        )

    parser = configparser.ConfigParser()
    parser.read(config_path, encoding="utf-8-sig")

    if CONFIG_SECTION not in parser:
        raise ValueError(f"配置文件缺少节: [{CONFIG_SECTION}]")

    section = parser[CONFIG_SECTION]
    template_raw = section.get("TEMPLATE_PATH", "").strip()
    inputs_raw = section.get("INPUT_PATHS", "").strip()
    title_map_raw = section.get("Title_Map", "").strip()
    task_assigner_map_raw = section.get("TaskAssigner_Map", "").strip()

    if not template_raw:
        raise ValueError("配置项 TEMPLATE_PATH 不能为空")
    if not inputs_raw:
        raise ValueError("配置项 INPUT_PATHS 不能为空")

    template_path = resolve_similar_excel_path(Path(template_raw))
    input_paths = [resolve_similar_excel_path(p) for p in parse_input_paths(inputs_raw)]
    if not input_paths:
        raise ValueError("INPUT_PATHS 未解析到有效文件路径")

    if not template_path.exists():
        raise FileNotFoundError(
            "TEMPLATE_PATH 文件不存在，请检查 [PATHS] 中 TEMPLATE_PATH 配置。\n"
            + build_path_hint(template_path)
        )
    missing_inputs = [p for p in input_paths if not p.exists()]
    existing_inputs = [p for p in input_paths if p.exists()]
    if not existing_inputs:
        hints = "\n\n".join(build_path_hint(p) for p in input_paths)
        raise FileNotFoundError(
            "INPUT_PATHS 中没有可用文件，请检查 [PATHS] 中 INPUT_PATHS 配置。\n"
            f"{hints}"
        )

    title_map = parse_title_map(title_map_raw)
    task_assigner_map = parse_task_assigner_map(task_assigner_map_raw)
    return template_path, existing_inputs, missing_inputs, title_map, task_assigner_map


def read_input_with_fallback(path: Path) -> pd.DataFrame:
    """
    Read preferred sheet 'Sheet0' when present, otherwise fallback to first sheet.
    """
    xls = pd.ExcelFile(path)
    preferred = "Sheet0"
    sheet_name = preferred if preferred in xls.sheet_names else xls.sheet_names[0]
    return pd.read_excel(path, sheet_name=sheet_name)


def infer_customer_type(customer_name: object) -> object:
    text = normalize(customer_name)
    if text == "":
        return pd.NA
    has_hq = "对公客户-总行级客户-" in text
    has_branch = "对公客户-分行级重点客户-" in text
    if has_hq:
        return "总行级客户"
    if has_branch:
        return "分行级重点客户"
    return pd.NA


def clean_customer_name(customer_name: object) -> object:
    text = normalize(customer_name)
    if text == "":
        return pd.NA
    text = text.replace("对公客户-总行级客户-", "")
    text = text.replace("对公客户-分行级重点客户-", "")
    return text.strip() or pd.NA

#  该程序将 总行部分的需求映射到李卫和张峰来处理
def infer_responder(org_name: object, responder_mapping: dict[str, str]) -> object:
    text = normalize(org_name)
    if text == "":
        return pd.NA
    if "中国" in text or "总行" in text:
        return "李卫、张峰"
    return responder_mapping.get(text, pd.NA)

# 如下代码设置日期格式字段为日期  步骤1
def parse_datetime_cell(value: object) -> object:
    if value is None or pd.isna(value):
        return pd.NA
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return pd.NA
    return ts.to_pydatetime()


def collect_duplicate_ids(df: pd.DataFrame, unique_key: str) -> list[str]:
    if unique_key not in df.columns:
        return []
    keys = df[unique_key].map(normalize)
    keys = keys[keys != ""]
    key_counts = keys.value_counts()
    duplicates = key_counts[key_counts > 1].index.tolist()
    return [str(x) for x in duplicates]


def emit_log(lines: list[str], log_path: Path) -> None:
    text = "\n".join(lines) + "\n"
    # [V10] 日志同时写入文件与标准输出，便于即时查看运行结果
    print(text, end="")
    with log_path.open("w", encoding="utf-8") as f:
        f.write(text)


def set_initial_view_to_template_tail(ws, original_template_rows: int) -> None:
    # [R12-01] 生成文件时设置初始可视区域：定位到“原模板数据倒数5行”的起始位置
    # 说明：模板数据从第2行开始（第1行为表头），若模板不足5行则回退到第2行。
    if original_template_rows <= 0:
        target_row = 2
    else:
        first_of_last_five_data_row = max(1, original_template_rows - 4)
        target_row = first_of_last_five_data_row + 1  # +1 对齐到 Excel 实际行号（含表头）

    target_cell = f"A{target_row}"
    ws.sheet_view.topLeftCell = target_cell
    if ws.sheet_view.selection:
        ws.sheet_view.selection[0].activeCell = target_cell
        ws.sheet_view.selection[0].sqref = target_cell


def main() -> None:
    # [V6->V8] 从固定配置文件读取路径、多输入文件、标题映射与回复人映射
    template_path, input_paths, missing_inputs, title_map, responder_mapping = load_config_from_ini(CONFIG_PATH)
    if missing_inputs:
        print("警告: 以下输入文件不存在，已自动跳过：")
        for p in missing_inputs:
            print(f" - {p}")

    # [V3] 读取模板页表头与已有数据，后续用于列对齐与增量合并
    template_df = pd.read_excel(template_path, sheet_name=TEMPLATE_SHEET)
    # [R12-02] 保留原模板数据量，用于设置输出文件打开时的初始定位行
    original_template_rows = len(template_df)
    template_headers = [normalize(c) for c in template_df.columns]
    template_header_set = set(template_headers)
    template_df.columns = template_headers

    merged_parts: list[pd.DataFrame] = []
    for p in input_paths:
        # [V6] 每个输入文件独立读取并统一映射后再汇总
        src = read_input_with_fallback(p)
        src.columns = [normalize(c) for c in src.columns]

        out = pd.DataFrame(index=src.index)
        # [V8] Title_Map 支持一对多映射（如 接收日期 -> [创建时间, 回复时间/ITM回函时间]）
        for src_col, target_cols in title_map.items():
            for target_col in target_cols:
                out[target_col] = src[src_col] if src_col in src.columns else pd.NA

        # V14:20260408 对一些固定的列做默认值填写
        out['事项来源'] = 'ITM'
        out['当前状态'] = '进行中'        
        
        # Fill non-mapped template columns with NA so append order stays aligned
        for col in template_headers:
            if col not in out.columns:
                out[col] = pd.NA
        out = out[template_headers]

        
        # Keep only columns that are truly in template
        out = out[[c for c in out.columns if c in template_header_set]]
        merged_parts.append(out)

    if not merged_parts:
        raise ValueError("没有可合并的输入数据，请检查 INPUT_PATHS 配置。")
    all_new_rows = pd.concat(merged_parts, ignore_index=True)

    unique_key = "工单编号/ITM编号"
    if unique_key not in template_headers:
        raise ValueError(f"模板页缺少唯一索引列: {unique_key}")

    # [R11-01] 先标准化唯一键，后续按“模板原顺序 + 新增排序追加”构建结果，避免混编
    template_base = template_df[template_headers].copy()
    template_base[unique_key] = template_base[unique_key].map(normalize)
    template_base = template_base[template_base[unique_key] != ""].reset_index(drop=True)

    new_rows = all_new_rows.copy()
    new_rows[unique_key] = new_rows[unique_key].map(normalize)
    new_rows = new_rows[new_rows[unique_key] != ""].reset_index(drop=True)

    # [R11-02] 统计去重前重复键（模板内、新数据内、模板与新数据交叉重复都纳入日志）
    duplicate_source = pd.concat([template_base, new_rows], ignore_index=True)
    duplicate_ids = collect_duplicate_ids(duplicate_source, unique_key)

    # [R11-03] 模板内若自身有重复，按最后一条保留，维持其在模板中的相对顺序
    template_base = template_base.drop_duplicates(subset=[unique_key], keep="last").reset_index(drop=True)

    # [R11-04] 新数据内先去重（同一工单保留最后一条），再按创建时间升序准备追加
    new_rows = new_rows.drop_duplicates(subset=[unique_key], keep="last").reset_index(drop=True)
    append_rows = new_rows[~new_rows[unique_key].isin(template_base[unique_key])].copy()
    if "创建时间" in append_rows.columns:
        append_rows["_sort_created_time"] = pd.to_datetime(append_rows["创建时间"], errors="coerce")
        append_rows = append_rows.sort_values(
            by=["_sort_created_time", unique_key], ascending=[True, True], na_position="last"
        ).drop(columns=["_sort_created_time"])
    append_rows = append_rows.reset_index(drop=True)

    # [R11-05] 最终结果为“模板原数据 + 新增排序数据”，保证新增记录统一追加在末尾
    combined = pd.concat([template_base, append_rows], ignore_index=True)

    customer_name_col = "客户名称"
    customer_type_col = "客户类型"
    if customer_name_col in combined.columns:
        if customer_type_col in combined.columns:
            # [V4] 仅当可识别客户类型时覆盖，避免误清空原模板已有值
            inferred = combined[customer_name_col].map(infer_customer_type)
            existing = combined[customer_type_col]
            combined[customer_type_col] = inferred.where(~inferred.isna(), existing)
        combined[customer_name_col] = combined[customer_name_col].map(clean_customer_name)

    creator_org_col = "创建人机构"
    responder_col = "回复人"
    if creator_org_col in combined.columns and responder_col in combined.columns:
        combined[responder_col] = combined[creator_org_col].map(
            lambda x: infer_responder(x, responder_mapping)
        )
    # 如下代码设置日期格式字段为日期  步骤2
    date_cols = ["创建时间", "回复时间/ITM回函时间"]
    for col in date_cols:
        if col in combined.columns:
            combined[col] = combined[col].map(parse_datetime_cell)

    # [V1/V3] 复制模板结构并回写去重后的结果到模板页
    output_path = build_output_path(template_path)
    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET]

    # Clear existing data rows but keep header row and formatting as much as possible.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    start_row = 2
    for i, row in combined.iterrows():
        for j, col in enumerate(template_headers, start=1):
            ws.cell(row=start_row + i, column=j, value=None if pd.isna(row[col]) else row[col])

    # 如下代码设置日期格式字段为日期  步骤3
    date_number_format = "yyyy-mm-dd hh:mm:ss"
    for col in date_cols:
        if col in template_headers:
            col_idx = template_headers.index(col) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = date_number_format
            # Ensure date-time text is fully visible: "YYYY-MM-DD HH:MM:SS"
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 21

    # [R12-03] 新增初始视图定位：打开文件时默认跳到原模板倒数5行（而非首行）
    set_initial_view_to_template_tail(ws, original_template_rows)

    wb.save(output_path)

    # [V9->V10] 输出运行日志：写入 .log 文件并同步打印到标准输出
    log_path = build_log_path(Path(__file__))
    log_lines: list[str] = [
        f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"模板文件: {template_path}",
        "输入文件:",
    ]
    log_lines.extend([f" - {p}" for p in input_paths])
    if missing_inputs:
        log_lines.append("缺失输入文件(已跳过):")
        log_lines.extend([f" - {p}" for p in missing_inputs])
    log_lines.extend(
        [
            f"输出文件: {output_path}",
            f"日志文件: {log_path}",
            f"导入原始记录数: {len(all_new_rows)}",
            f"新增记录数(追加到模板末尾): {len(append_rows)}",
            f"去重后总记录数: {len(combined)}",
            f"重复记录数(按{unique_key}): {len(duplicate_ids)}",
            f"重复{unique_key}值:",
        ]
    )
    if duplicate_ids:
        log_lines.extend([f" - {key}" for key in duplicate_ids])
    else:
        log_lines.append(" - 无")
    emit_log(log_lines, log_path)


# 解决 pyinstaller 打包后找不到配置文件的问题
def get_base_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent

if __name__ == "__main__":
    base_dir = get_base_dir()
    CONFIG_PATH = base_dir / 'merge_requirement_to_template.ini'
    print(f"CONFIG_PATH: {CONFIG_PATH}")

    if not CONFIG_PATH.exists():
        print(f"未找到配置文件: {CONFIG_PATH}")
        sys.exit(1)

    try:
        main()
    except Exception as exc:
        print("程序执行失败，请检查以下错误信息：", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        sys.exit(1)